"""공공데이터(전국 병의원 및 약국 현황) CSV에서 병원 카탈로그를 만든다.

이 데이터는 여러 CSV로 나뉜다:
- 기본정보 CSV: 요양기관명, 종별코드명, (병원)홈페이지, 암호화요양기호, 시도 등
- 시설정보 CSV: 암호화요양기호, 총병상수 (병상수)

두 파일을 암호화요양기호로 조인해 "종합병원/상급종합병원 & 병상수>=N"을 필터링한다.
인코딩은 CP949(EUC-KR)인 경우가 많아 자동 감지한다. 컬럼명이 판마다 다를 수 있어
헤더 부분일치로 관대하게 찾는다.
"""
from __future__ import annotations

import csv
import io
from pathlib import Path

# 공공데이터의 종별코드명은 상급종합병원을 "상급종합"으로 표기한다("병원" 없음).
DEFAULT_TYPES = ("종합병원", "상급종합", "상급종합병원")

# 헤더 부분일치 후보(왼쪽 우선).
_NAME_KEYS = ("요양기관명", "병원명", "기관명")
_TYPE_KEYS = ("종별코드명", "종별명", "종별")
_HOME_KEYS = ("병원홈페이지", "홈페이지", "url")
_CODE_KEYS = ("암호화요양기호", "암호화요양기관기호", "요양기호", "요양기관기호")
_REGION_KEYS = ("시도코드명", "시도명", "시도", "주소")
# 시설정보 CSV/XLSX에는 단일 '총병상수'가 없고 병상 종류별 컬럼('...병상수')으로 나뉜다.
# 입원 병상만 합산하기 위해 아래 종류는 제외한다(입원 병상이 아님).
_BED_COL_HINT = "병상수"
_BED_EXCLUDE = ("분만실", "수술실", "응급실", "물리치료실")


def _decode(path: str | Path) -> str:
    data = Path(path).read_bytes()
    for enc in ("utf-8-sig", "cp949", "euc-kr", "utf-8"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def _read_xlsx(path: str | Path) -> tuple[list[str], list[dict[str, str]]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    headers = [(str(h).strip() if h is not None else "") for h in next(it)]
    rows: list[dict[str, str]] = []
    for r in it:
        rows.append(
            {h: ("" if v is None else str(v)).strip() for h, v in zip(headers, r)}
        )
    wb.close()
    return headers, rows


def _read_rows(path: str | Path) -> tuple[list[str], list[dict[str, str]]]:
    if Path(path).suffix.lower() in (".xlsx", ".xlsm"):
        return _read_xlsx(path)
    reader = csv.DictReader(io.StringIO(_decode(path)))
    headers = [h.strip() for h in (reader.fieldnames or [])]
    rows = [{(k.strip() if k else k): (v or "").strip() for k, v in r.items()} for r in reader]
    return headers, rows


def _bed_columns(headers: list[str]) -> list[str]:
    return [
        h for h in headers if _BED_COL_HINT in h and not any(x in h for x in _BED_EXCLUDE)
    ]


def _find_col(headers: list[str], keys: tuple[str, ...]) -> str | None:
    for key in keys:
        for h in headers:
            if key in h:
                return h
    return None


def _to_int(s: str) -> int | None:
    digits = "".join(ch for ch in (s or "") if ch.isdigit())
    return int(digits) if digits else None


def load_bed_counts(facility_path: str | Path) -> dict[str, int]:
    """시설정보에서 암호화요양기호 -> 입원 병상 합계 매핑을 만든다.

    단일 '총병상수' 컬럼이 없으므로 '...병상수' 컬럼들(분만실/수술실/응급실/물리치료실 제외)을
    행마다 합산한다.
    """
    headers, rows = _read_rows(facility_path)
    code_col = _find_col(headers, _CODE_KEYS)
    bed_cols = _bed_columns(headers)
    if not code_col or not bed_cols:
        raise ValueError(f"시설 파일에서 코드/병상 컬럼을 못 찾음: headers={headers}")
    out: dict[str, int] = {}
    for r in rows:
        code = r.get(code_col)
        if not code:
            continue
        out[code] = sum((_to_int(r.get(c, "")) or 0) for c in bed_cols)
    return out


def load_catalog(
    basic_path: str | Path,
    *,
    facility_path: str | Path | None = None,
    types: tuple[str, ...] = DEFAULT_TYPES,
    min_beds: int | None = 300,
) -> list[dict]:
    """기본정보 CSV(+선택 시설정보 CSV)에서 필터링된 병원 카탈로그를 반환한다.

    각 항목: {name, type, region, homepage, beds, code}.
    facility_path가 없으면 병상수 필터는 적용하지 않고 beds=None으로 둔다.
    """
    headers, rows = _read_rows(basic_path)
    name_col = _find_col(headers, _NAME_KEYS)
    type_col = _find_col(headers, _TYPE_KEYS)
    if not name_col or not type_col:
        raise ValueError(f"기본 CSV에서 기관명/종별 컬럼을 못 찾음: headers={headers}")
    home_col = _find_col(headers, _HOME_KEYS)
    code_col = _find_col(headers, _CODE_KEYS)
    region_col = _find_col(headers, _REGION_KEYS)

    beds_by_code: dict[str, int] = {}
    if facility_path is not None:
        beds_by_code = load_bed_counts(facility_path)

    catalog: list[dict] = []
    for r in rows:
        if r.get(type_col) not in types:
            continue
        code = r.get(code_col) if code_col else None
        beds = beds_by_code.get(code) if code else None
        if min_beds is not None and facility_path is not None:
            if beds is None or beds < min_beds:
                continue
        homepage = r.get(home_col, "") if home_col else ""
        if homepage.lower() == "none":  # 홈페이지 없음이 문자열 'None'으로 저장됨
            homepage = ""
        catalog.append(
            {
                "name": r.get(name_col, ""),
                "type": r.get(type_col, ""),
                "region": r.get(region_col, "") if region_col else "",
                "homepage": homepage,
                "beds": beds,
                "code": code,
            }
        )
    return catalog
